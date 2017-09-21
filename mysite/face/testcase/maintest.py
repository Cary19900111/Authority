#！/usr/bin/env python
#coding:utf-8
__author__ = 'caryr'
import unittest,os,HTMLTestRunner,sys
from .ci.test2basedata import Testjiuyelv
from .ci.test1uisixindicator import TestUi
from .jp.testpublishtask import TestPublish
from datetime import datetime
from .ci.common.ParaTestCase import ParametrizedTestCase
from openpyxl import load_workbook

def get_ci_information():
    uni_all=[]
    abs_path = os.getcwd()
    file_path  = abs_path+r"/face/testcase/ci/uploadfile/ci.xlsx"
    file = load_workbook(file_path)
    sheet_half = file.get_sheet_by_name('Sheet1')
    for rx in range(sheet_half.max_row-1):
        uni = {}
        uni['account']  = sheet_half.cell(row=rx + 2, column=1).value
        uni['password'] = sheet_half.cell(row=rx + 2, column=2).value
        uni['code'] = sheet_half.cell(row=rx + 2, column=3).value
        uni['id'] = sheet_half.cell(row=rx + 2, column=4).value
        uni['db'] = sheet_half.cell(row=rx + 2, column=5).value
        uni['year'] = sheet_half.cell(row=rx + 2, column=6).value
        uni['product']=sheet_half.cell(row=rx + 2, column=7).value
        uni['level'] = sheet_half.cell(row=rx + 2, column=8).value #level=1 bachelor
        uni_all.append(uni)
    file.save(file_path)
    return uni_all

def get_jp_information():
    uni_all=[]
    abs_path = os.getcwd()
    file_path  = abs_path+r"/face/testcase/jp/uploadfile/jp.xlsx"
    file = load_workbook(file_path)
    sheet_half = file.get_sheet_by_name('Sheet1')
    #range 左闭右开
    for rx in range(sheet_half.max_row-1):
        uni = {}
        uni['ip']  = sheet_half.cell(row=rx + 2, column=1).value
        uni['account'] = sheet_half.cell(row=rx + 2, column=2).value
        uni['password'] = sheet_half.cell(row=rx + 2, column=3).value
        uni['name'] = sheet_half.cell(row=rx + 2, column=4).value
        uni_all.append(uni)
    file.save(file_path)
    return uni_all

def test_ci_all_case(filename):
    try:
        university_list = get_ci_information()
        abs_path = os.path.abspath('.')
        test_dir = abs_path + r'\face\testcase\ci'
        #test_dir = abs_path + r'\ci'
        HtmlFile = r'{abs_path}\face\testcase\result\{filename}'.format(abs_path=abs_path, filename=filename)
        # testunit = unittest.TestSuite()
        # discover = unittest.defaultTestLoader.discover(test_dir, pattern='test*.py',
        #                                                top_level_dir=None)  # 返回多个suite的列表，一个文件是一个suite
        # for test_suite in discover:  # 得到一个suite，一个文件
        #     for test_case in test_suite:  # 得到一个case,一个case是，里面的一个函数
        #         pass
        #     testunit.addTests(test_case)
        fp = open(HtmlFile, "wb")
        for university in university_list:
            runner = HTMLTestRunner.HTMLTestRunner(stream=fp, title=str(university['code']), description=u"测试报告")
            suite = unittest.TestSuite()
            # addTest既可以是suite，也可以是tests
            #suite.addTest(ParametrizedTestCase.parametrize(Testjiuyelv, param=university))
            suite.addTest(ParametrizedTestCase.parametrize(TestUi, param=university))
            runner.run(suite)
        fp.close()
    except Exception as err:
        print(err)

def test_jp_all_case(filename):
    try:
        university_list = get_jp_information()
        abs_path = os.path.abspath('.')
        test_dir = abs_path + r'\face\testcase\jp'
        #test_dir = abs_path + r'\ci'
        HtmlFile = r'{abs_path}\face\testcase\result\{filename}'.format(abs_path=abs_path, filename=filename)
        # testunit = unittest.TestSuite()
        # discover = unittest.defaultTestLoader.discover(test_dir, pattern='test*.py',
        #                                                top_level_dir=None)  # 返回多个suite的列表，一个文件是一个suite
        # for test_suite in discover:  # 得到一个suite，一个文件
        #     for test_case in test_suite:  # 得到一个case,一个case是，里面的一个函数
        #         pass
        #     testunit.addTests(test_case)
        fp = open(HtmlFile, "wb")
        for university in university_list:
            runner = HTMLTestRunner.HTMLTestRunner(stream=fp, title=str(university['name']), description=u"测试报告")
            suite = unittest.TestSuite()
            # addTest既可以是suite，也可以是tests
            #suite.addTest(ParametrizedTestCase.parametrize(Testjiuyelv, param=university))
            suite.addTest(ParametrizedTestCase.parametrize(TestPublish, param=university))
            runner.run(suite)
        fp.close()
    except Exception as err:
        print(err)

if __name__ == '__main__':
    test_ci_all_case("123")